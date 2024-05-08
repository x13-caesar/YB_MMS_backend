import glob
import os
from datetime import datetime, timedelta

from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

from app import models
from app.routers.business import get_company_info, get_kaipiao_info
from app.routers.specification import get_component_by_specification_id


def generate_formatted_instock_form(sheet, form: models.InstockForm, db):
    '''

    :param sheet:
    :return:
    '''
    n_items = len(form.instock_item)
    vendor = form.vendor
    kaipiao_info = get_kaipiao_info()
    company_info = get_company_info()

    # styles
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    title_alignment = Alignment(horizontal="center", vertical="center")
    general_alignment = Alignment(horizontal="general", vertical="center")

    # merging and formatting
    row_comment = 6 + n_items + 2  # "说明"所在行，多 +1 因为有合计行
    row_manager = row_comment + 1
    row_confirmation = row_comment + 2
    row_company_info = row_confirmation + 2
    area_to_merge = [
        "A1:I1",
        "A2:I2",
        "D3:E3",
        "D4:E4",
        "D5:E5",
        "F3:G3",
        "F4:G4",
        "F5:G5",
        "H3:I3",
        "H4:I4",
        "H5:I5",
        f"B{row_comment}:I{row_comment}",
        f"B{row_confirmation}:I{row_confirmation}",
    ]
    # 正文合并单元格
    for area in area_to_merge:
        sheet.merge_cells(area)
    # 公司信息合并单元格
    sheet.merge_cells(f"A{row_company_info - 1}:I{row_company_info - 1}")
    for r in range(row_company_info, row_company_info + 7):
        sheet.merge_cells(f"A{r}:E{r}")
        sheet.merge_cells(f"F{r}:I{r}")

    # 固定文字
    sheet["A2"].value = "采购订单"
    sheet["A3"].value = "供方传真"
    sheet["C3"].value = "供方编号"
    sheet["F3"].value = "供方名称"
    sheet["A4"].value = "供方联系人"
    sheet["C4"].value = "供方联系人电话"
    sheet["F4"].value = "供方电邮"
    sheet["A5"].value = "采购订单系统编号"
    sheet["C5"].value = "下单日期"
    sheet["F5"].value = "供应商地址"
    sheet["A6"].value = "物料编号"
    sheet["B6"].value = "物料名称"
    sheet["C6"].value = "型号"
    sheet["D6"].value = "采购数量"
    sheet["E6"].value = "单位"
    sheet["F6"].value = "单价"
    sheet["G6"].value = "到货时间"
    sheet["H6"].value = "包装数量"
    sheet["I6"].value = "备注"
    sheet[f"A{str(row_comment - 1)}"].value = "合计"
    sheet[f"A{str(row_comment)}"].value = "说明"
    sheet[f"B{str(row_comment)}"].value = "供方在接到本单后，应在8小时内传真或邮件回复我司供应部，若未在规定时间内回复，默认为接受并履行订购单内全部内容。"
    sheet.cell(row=row_manager, column=1).value = "采购员"
    sheet.cell(row=row_manager, column=3).value = "采购员电话"
    sheet.cell(row=row_manager, column=6).value = "审核"
    sheet.cell(row=row_manager, column=8).value = "批准人"
    sheet.cell(row=row_confirmation, column=1).value = "供方确认"
    sheet.cell(row=row_company_info, column=1).value = "供方确认电话"

    # 基本信息
    sheet["A1"].value = company_info["公司名称"]
    sheet.cell(row=3, column=2).value = vendor.fax
    sheet.cell(row=3, column=4).value = str(vendor.id).rjust(3, '0')
    sheet["H3"].value = vendor.company
    sheet.cell(row=4, column=2).value = vendor.name
    sheet.cell(row=4, column=4).value = vendor.contact
    sheet["H4"].value = vendor.email
    sheet.cell(row=5, column=2).value = form.display_form_id
    sheet.cell(row=5, column=4).value = form.create_time.date()
    sheet["H5"].value = vendor.address

    # 公司及开票信息
    sheet[f"A{row_company_info}"].value = '公司基本信息及联系方式：'
    sheet[f"F{row_company_info}"].value = '开票资料：'
    for r, kv in enumerate(company_info.items()):
        sheet[f"A{row_company_info + 1 + r}"].value = f"{str(kv[0])}：{str(kv[1])}"
    for r, kv in enumerate(kaipiao_info.items()):
        sheet[f"F{row_company_info + 1 + r}"].value = f"{str(kv[0])}：{str(kv[1])}"

    # 采购内容
    total_qty, total_price = 0, 0
    for idx, item in enumerate(form.instock_item):
        row_item = 7 + idx
        target_component = get_component_by_specification_id(item.specification_id, db)
        for spec in target_component.specification:
            if spec.id == item.specification_id:
                target_specification = spec
        sheet.cell(row=row_item, column=1).value = item.specification_id  # 物料编号
        sheet.cell(row=row_item, column=2).value = target_component.name  # 物料名称
        sheet.cell(row=row_item, column=3).value = target_component.model  # 型号
        sheet.cell(row=row_item, column=4).value = item.order_quantity  # 采购数量
        total_qty += item.order_quantity  # 计入数量合计
        sheet.cell(row=row_item, column=5).value = target_component.as_unit  # 单位
        # sheet.cell(row=row_item, column=6).value = item.unit_cost  # 单价
        total_price += item.order_quantity * item.unit_cost  # 计入价格合计
        sheet.cell(row=row_item, column=7).value = item.instock_date  # 到货时间
        sheet.cell(
            row=row_item, column=8
        ).value = target_specification.unit_amount  # 包装数量
        sheet.cell(row=row_item, column=9).value = item.notice  # 备注
    # 合计
    sheet.cell(row=row_comment - 1, column=4).value = total_qty
    # sheet.cell(row=row_comment - 1, column=6).value = total_price

    for row in sheet.iter_rows():
        for cell in row:
            # 标题行
            if cell.row <= 2:
                cell.border = border
                cell.alignment = title_alignment
                cell.font = cell.font.copy(bold=True)
                sheet.row_dimensions[cell.row].height = 30
            # 采购物料行，需要显示全部备注文字，所以高度90
            elif 6 < cell.row < row_comment:
                cell.border = border
                sheet.row_dimensions[cell.row].height = 60
                cell.alignment = general_alignment
                cell.alignment = cell.alignment.copy(wrapText=True)
            # 其他行30
            elif cell.row < row_company_info - 1:
                cell.border = border
                cell.alignment = general_alignment
                sheet.row_dimensions[cell.row].height = 30
            else:
                cell.alignment = general_alignment
                sheet.row_dimensions[cell.row].height = 30
            # 设置单元格宽度为40
            try:
                sheet.column_dimensions[cell.column_letter].width = 20
            except Exception as e:
                print(e)
    # 备注列单独加宽
    sheet.column_dimensions["I"].width = 40

    save_as_filename = f"{form.form_id}_{vendor.company}.xlsx"

    return save_as_filename


def wipe_old_files(working_directory=os.getcwd()):
    # Get a list of .xlsx files in the current working directory
    xlsx_files = glob.glob(os.path.join(working_directory, "*.xlsx"))
    # Remove each .xlsx file
    for file_path in xlsx_files:
        time_created = datetime.fromtimestamp(os.path.getctime(file_path))
        # remove it if it's created before today
        one_day_ago = datetime.now() - timedelta(days=1)
        if time_created <= one_day_ago:
            print(f"Deleting {file_path}")
            os.remove(file_path)
